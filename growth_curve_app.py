import streamlit as st
import openpyxl
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import os
import json

# ══════════════════════════════════════════════════════════════════
# 성장곡선 검증 Streamlit 웹 앱
# ══════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="예측 기준 매출 검증 프로그램",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── 스타일 ────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        font-size: 2rem;
        font-weight: bold;
        color: #1A3A5C;
        padding: 0.5rem 0;
    }
    .sub-header {
        font-size: 1rem;
        color: #5D6D7E;
        margin-bottom: 1rem;
    }
    .metric-card {
        background: #EBF5FB;
        border-radius: 8px;
        padding: 1rem;
        border-left: 4px solid #2471A3;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
</style>
""", unsafe_allow_html=True)

# ── 파일 경로 ─────────────────────────────────────────────────────
FILE_PATH = os.path.join(os.path.dirname(__file__),
                         "260618 경쟁사유무에 따른 성장곡선(공유)V4(시즌미반영성장곡선).xlsx")
ADDED_STORES_PATH = os.path.join(os.path.dirname(__file__),
                                  "added_stores.json")
TARGET_PATH = os.path.join(os.path.dirname(__file__), "타겟수요.xlsx")


# ── 추가 매장 데이터 저장/로드 ────────────────────────────────────
def load_added_stores():
    if os.path.exists(ADDED_STORES_PATH):
        with open(ADDED_STORES_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def save_added_stores(stores_list):
    with open(ADDED_STORES_PATH, 'w', encoding='utf-8') as f:
        json.dump(stores_list, f, ensure_ascii=False, indent=2)


# ── 타겟수요 로드 ────────────────────────────────────────────────
@st.cache_data
def load_target_data():
    """타겟수요.xlsx에서 매장별 타겟수요를 로드"""
    if not os.path.exists(TARGET_PATH):
        return {}
    wb = openpyxl.load_workbook(TARGET_PATH, data_only=True, read_only=True)
    ws = wb.active
    targets = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        name = row[0]
        target = row[1]
        if name and target:
            targets[str(name).strip()] = float(target)
    wb.close()
    return targets


# ── 회귀밴드 계산 ────────────────────────────────────────────────
REGRESSION_COEF = 5908.958069
REGRESSION_INTERCEPT = 2948754.80292545
REGRESSION_BAND = 1300233.89703197


def get_regression_band(target_demand):
    """타겟수요로 예상매출/상한/하한 계산"""
    pred = target_demand * REGRESSION_COEF + REGRESSION_INTERCEPT
    return {
        'predicted': pred,
        'upper': pred + REGRESSION_BAND,
        'lower': pred - REGRESSION_BAND
    }


def match_target(store_name, target_data):
    """매장명에서 지점명을 추출하여 타겟수요 매칭 (#제거, 괄호 제거)"""
    # (1143) #장항점 → 장항점
    parts = store_name.split(') ')
    short = parts[1].strip() if len(parts) > 1 else store_name.strip()
    short = short.lstrip('#')
    # 타겟수요에서도 # 제거하고 비교
    for key, val in target_data.items():
        if key.lstrip('#') == short:
            return val
    return None


# ── 엑셀 데이터 로드 (캐시) ───────────────────────────────────────
@st.cache_data
def load_excel_data():
    """V3 엑셀에서 곡선지수와 매장 매출 데이터를 로드"""
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)

    # 1. 가중평균 곡선지수 (요약_성장곡선(시즌미반영) 시트, 6열: 가중평균(0~2))
    ws_curve = wb['요약_성장곡선(시즌미반영)']
    curve_index = {}
    for row in ws_curve.iter_rows(min_row=3, max_row=150, max_col=10,
                                   values_only=True):
        m = row[0]   # 오픈후개월(m)
        w_avg = row[5]  # 가중평균(0~2)
        if m is not None and w_avg is not None and w_avg != 0:
            curve_index[int(m)] = w_avg

    # 2. RAW_매출A: 세로 형태 → 매장별 월매출 리스트로 변환
    ws_raw = wb['RAW_매출A']
    from collections import defaultdict
    store_data = defaultdict(lambda: {'open_date': None, 'monthly': {}})

    for row in ws_raw.iter_rows(min_row=2, max_col=10, values_only=True):
        date_val = row[0]      # 해당월
        store_name = row[1]    # 지점키
        open_date = row[3]     # 개점일
        sales = row[6]         # 세탁건조매출

        if store_name is None or date_val is None:
            continue

        if store_data[store_name]['open_date'] is None and open_date:
            store_data[store_name]['open_date'] = open_date

        if sales is not None and sales > 0:
            store_data[store_name]['monthly'][date_val] = sales

    # 매장별로 개점일 기준 월차(m0, m1, ...) 순서로 매출 정렬
    stores = []
    # 장항점 통합: (1009)#진시스장항점, (1143)#진시스장항점, (1143)#장항점 → 장항점
    janghang_keys = ['(1009) #진시스장항점', '(1143) #진시스장항점', '(1143) #장항점']
    janghang_merged = {'open_date': None, 'monthly': {}}
    # 트리하우스 제외
    exclude_keys = ['(0008) 트리하우스점']

    for name, data in store_data.items():
        if not data['monthly'] or not data['open_date']:
            continue
        if name in exclude_keys:
            continue
        if name in janghang_keys:
            # 장항점 데이터 합치기 (날짜 중복 시 덮어쓰기)
            if janghang_merged['open_date'] is None:
                janghang_merged['open_date'] = data['open_date']
            elif data['open_date'] < janghang_merged['open_date']:
                janghang_merged['open_date'] = data['open_date']
            janghang_merged['monthly'].update(data['monthly'])
            continue

        # 날짜순 정렬
        sorted_months = sorted(data['monthly'].keys())
        sales_list = [data['monthly'][m] for m in sorted_months]

        if len(sales_list) >= 4:  # 최소 데이터 필요
            stores.append({'name': name, 'sales': sales_list})

    # 장항점 통합 데이터 추가
    if janghang_merged['monthly']:
        sorted_months = sorted(janghang_merged['monthly'].keys())
        sales_list = [janghang_merged['monthly'][m] for m in sorted_months]
        if len(sales_list) >= 4:
            stores.append({'name': '(1143) #장항점', 'sales': sales_list})

    wb.close()
    return curve_index, stores


# ── 검증 함수 ─────────────────────────────────────────────────────
def get_base_revenue(store, curve_index, method):
    """단일 방식(A/B/C)의 기준매출 계산. 실패 시 None 반환."""
    sales = store['sales']

    start_m = {'A': 1, 'B': 2, 'C': 3}[method]
    end_m = 3

    base_estimates = []
    for m in range(start_m, end_m + 1):
        if m >= len(sales) or sales[m] is None or sales[m] <= 0:
            continue
        if m not in curve_index:
            continue
        base_est = sales[m] / (curve_index[m] / 100)
        base_estimates.append(base_est)

    if not base_estimates:
        return None
    return np.mean(base_estimates)


def get_store_max_month(store):
    """매장의 최대 유효 월차 반환 (0-indexed)"""
    sales = store['sales']
    max_m = 0
    for i, s in enumerate(sales):
        if s is not None and s > 0:
            max_m = i
    return max_m


# 전체 방식 목록
ALL_METHODS = ['A', 'B', 'C', 'AB', 'AC', 'BC', 'ABC']
METHOD_LABELS = {
    'A': 'm1,m2,m3 사용', 'B': 'm2,m3 사용', 'C': 'm3 사용',
    'D': 'm4 사용', 'E': 'm5 사용', 'F': 'm6 사용', 'G': 'm7 사용', 'H': 'm8 사용',
    'AB': 'A+B 평균', 'AC': 'A+C 평균', 'BC': 'B+C 평균', 'ABC': 'A+B+C 평균'
}


def validate_store(store, curve_index, method='A', version='v1'):
    """
    method: 'A','B','C','D','E' = 단일, 'AB','AC','BC','ABC' = 복합(기준매출 평균)
    version: 'v1' = m1~m3 역산, m4~m9 검증
             'v2' = m1~m4 역산, m5~m9 검증
             'v3' = m1~m5 역산, m6~m9 검증
    """
    sales = store['sales']

    # 버전별 설정 (역산 범위만 달라짐)
    if version == 'v1':
        end_m = 3
    elif version == 'v2':
        end_m = 4
    elif version == 'v3':
        end_m = 5
    elif version == 'v4':
        end_m = 6
    elif version == 'v5':
        end_m = 7
    elif version == 'v6':
        end_m = 8
    else:
        end_m = 3

    # 기준매출 계산
    if method in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'):
        start_m_map = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8}
        start_m = start_m_map.get(method)
        if start_m is None or start_m > end_m:
            return None
        base_estimates = []
        for m in range(start_m, end_m + 1):
            if m >= len(sales) or sales[m] is None or sales[m] <= 0:
                continue
            if m not in curve_index:
                continue
            base_estimates.append(sales[m] / (curve_index[m] / 100))
        if not base_estimates:
            return None
        base_revenue = np.mean(base_estimates)
    elif method == 'AB':
        a = _calc_base(store, curve_index, 'A', end_m)
        b = _calc_base(store, curve_index, 'B', end_m)
        if a is None or b is None:
            return None
        base_revenue = (a + b) / 2
    elif method == 'AC':
        a = _calc_base(store, curve_index, 'A', end_m)
        c = _calc_base(store, curve_index, 'C', end_m)
        if a is None or c is None:
            return None
        base_revenue = (a + c) / 2
    elif method == 'BC':
        b = _calc_base(store, curve_index, 'B', end_m)
        c = _calc_base(store, curve_index, 'C', end_m)
        if b is None or c is None:
            return None
        base_revenue = (b + c) / 2
    elif method == 'ABC':
        a = _calc_base(store, curve_index, 'A', end_m)
        b = _calc_base(store, curve_index, 'B', end_m)
        c = _calc_base(store, curve_index, 'C', end_m)
        if a is None or b is None or c is None:
            return None
        base_revenue = (a + b + c) / 3
    else:
        return None

    # 검증: 항상 m4~m9 실제 평균과 비교
    if len(sales) >= 10:
        actual_m4_m9 = sales[4:10]
        if all(v is not None and v > 0 for v in actual_m4_m9):
            predicted = []
            for m in range(4, 10):
                if m in curve_index:
                    predicted.append(base_revenue * (curve_index[m] / 100))
                else:
                    predicted = None
                    break

            if predicted:
                errors = [(p - a) / a * 100 for p, a in zip(predicted, actual_m4_m9)]
                actual_avg = np.mean(actual_m4_m9)
                avg_error = (base_revenue - actual_avg) / actual_avg * 100
                return {
                    'base_revenue': base_revenue,
                    'predicted': predicted,
                    'actual': actual_m4_m9,
                    'errors': errors,
                    'avg_error': avg_error,
                    'verify_start': 4,
                    'verify_end': 10
                }

    # m4~m9 부족 → 기준매출만 반환
    return {
        'base_revenue': base_revenue,
        'predicted': None,
        'actual': None,
        'errors': None,
        'avg_error': None,
        'verify_start': 4,
        'verify_end': 10
    }


def _calc_base(store, curve_index, method, end_m):
    """복합 방식용 내부 기준매출 계산"""
    sales = store['sales']
    start_m_map = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8}
    start_m = start_m_map.get(method)
    if start_m is None or start_m > end_m:
        return None
    base_estimates = []
    for m in range(start_m, end_m + 1):
        if m >= len(sales) or sales[m] is None or sales[m] <= 0:
            continue
        if m not in curve_index:
            continue
        base_estimates.append(sales[m] / (curve_index[m] / 100))
    if not base_estimates:
        return None
    return np.mean(base_estimates)


# ── 전체 성장곡선 예측 (임의 개월까지) ────────────────────────────
def predict_growth_curve(base_revenue, curve_index, months=48):
    """기준매출과 곡선지수로 월별 예측매출 생성"""
    predictions = {}
    for m in range(1, months + 1):
        if m in curve_index:
            predictions[m] = base_revenue * (curve_index[m] / 100)
    return predictions


# ══════════════════════════════════════════════════════════════════
# 메인 앱
# ══════════════════════════════════════════════════════════════════
def main():
    st.markdown('<div class="main-header">📈 예측 기준 매출 검증 프로그램</div>',
                unsafe_allow_html=True)
    st.markdown('<div class="sub-header">'
                '예측 기준 매출 vs 실제 기준 매출 오차율 검증 | 곡선지수 : 그룹 0~2 매장 수 가중 평균</div>',
                unsafe_allow_html=True)

    # 데이터 로드
    if not os.path.exists(FILE_PATH):
        st.error(f"엑셀 파일을 찾을 수 없습니다: {FILE_PATH}")
        st.info("파일 경로를 확인하거나 엑셀 파일을 업로드해주세요.")
        return

    curve_index, stores = load_excel_data()

    # 추가 매장 데이터 로드
    added_stores = load_added_stores()
    all_stores = stores + added_stores

    # 사이드바
    with st.sidebar:
        st.header("⚙️ 설정")
        st.markdown("---")
        st.subheader("📊 데이터 현황")
        st.metric("엑셀 매장 수", len(stores))
        st.metric("추가 매장 수", len(added_stores))
        st.metric("곡선지수 월차 수", len(curve_index))

    # 탭 구성
    tab1, tab2, tab5 = st.tabs([
        "📊 전체 검증 결과", "🔍 개별 매장 조회",
        "📉 그룹 별 표준편차 분석"
    ])

    # ═══════════════════════════════════════════════════════════════
    # 탭 1: 전체 검증 결과
    # ═══════════════════════════════════════════════════════════════
    with tab1:
        st.subheader("🏪 전체 매장 검증 결과")

        # 버전 선택
        version_labels = {
            'v1': 'V1: m1~m3 데이터 사용',
            'v2': 'V2: m1~m4 데이터 사용',
            'v3': 'V3: m1~m5 데이터 사용',
            'v4': 'V4: m1~m6 데이터 사용',
            'v5': 'V5: m1~m7 데이터 사용',
            'v6': 'V6: m1~m8 데이터 사용'
        }
        selected_version = st.radio(
            "검증 버전 선택", ['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
            format_func=lambda x: version_labels[x],
            horizontal=True, key="version_sel"
        )

        # 버전별 방식 목록 (모든 버전 동일: A,B,C + 복합)
        version_methods = ALL_METHODS  # A,B,C,AB,AC,BC,ABC

        # 버전별 방식 라벨 (동적, 물결표시)
        end_m_map = {'v1': 3, 'v2': 4, 'v3': 5, 'v4': 6, 'v5': 7, 'v6': 8}
        ver_end = end_m_map[selected_version]
        ver_labels = {}
        for m in version_methods:
            start_m = {'A': 1, 'B': 2, 'C': 3}.get(m)
            if start_m and start_m <= ver_end:
                if start_m == ver_end:
                    ver_labels[m] = f'm{start_m} 사용'
                else:
                    ver_labels[m] = f'm{start_m}~m{ver_end} 사용'
            else:
                ver_labels[m] = METHOD_LABELS.get(m, '')

        # 상세 표에 사용할 방식 선택 (버전별 라벨)
        def get_method_label(method, ver):
            end_m = {'v1': 3, 'v2': 4, 'v3': 5, 'v4': 6, 'v5': 7, 'v6': 8}[ver]
            start_m = {'A': 1, 'B': 2, 'C': 3}.get(method)
            if start_m and start_m <= end_m:
                if start_m == end_m:
                    return f"{method}: m{start_m} 사용"
                else:
                    return f"{method}: m{start_m}~m{end_m} 사용"
            return f"{method}: {METHOD_LABELS.get(method, '')}"

        # 방식별 검증 실행
        results_all = {}
        for method in version_methods:
            results_all[method] = []
            for store in all_stores:
                result = validate_store(store, curve_index, method, selected_version)
                if result is not None and result['avg_error'] is not None:
                    results_all[method].append({
                        'name': store['name'], **result
                    })

        # 신뢰도 있는 평균 (트리밍 평균: 상하 10% 제거)
        def trimmed_mean(values, trim_pct=0.1):
            """상하 trim_pct% 제거 후 평균 (이상치 영향 최소화)"""
            if not values:
                return 0
            sorted_vals = sorted(values)
            n = len(sorted_vals)
            trim_n = int(n * trim_pct)
            if trim_n > 0:
                trimmed = sorted_vals[trim_n:-trim_n]
            else:
                trimmed = sorted_vals
            return np.mean(trimmed) if trimmed else np.mean(sorted_vals)

        # 최적 방식 찾기 (트리밍평균 → 표준편차 → 단순평균 순으로 비교)
        valid_methods = [m for m in version_methods if results_all.get(m)]
        best_overall = None
        if valid_methods:
            def method_score(m):
                errs = [abs(r['avg_error']) for r in results_all[m]]
                tm = round(trimmed_mean(errs), 2)
                std = np.std(errs)
                simple = np.mean(errs)
                return (tm, std, simple)
            best_overall = min(valid_methods, key=method_score)

        # 요약 카드 (모든 버전 3+4 레이아웃)
        summary_html = '<div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin: 1rem 0;">'
        first_row = ['A', 'B', 'C']
        second_row = ['AB', 'AC', 'BC', 'ABC']

        for method in first_row:
            errs = [abs(r['avg_error']) for r in results_all[method]]
            if errs:
                avg_abs_err = trimmed_mean(errs)
                std_err = np.std(errs)
                bias = np.mean([r['avg_error'] for r in results_all[method]])
                n_stores = len(errs)

                # 색상
                if method == best_overall:
                    border_color = '#2ECC71'
                    bg_color = '#EAFAF1'
                    badge = '<span style="background: #2ECC71; color: white; font-size: 0.6rem; padding: 2px 8px; border-radius: 10px;">★ 최적</span>'
                elif avg_abs_err <= 12:
                    border_color = '#3498DB'
                    bg_color = '#EBF5FB'
                    badge = ''
                elif avg_abs_err <= 18:
                    border_color = '#F39C12'
                    bg_color = '#FEF9E7'
                    badge = ''
                else:
                    border_color = '#E74C3C'
                    bg_color = '#FDEDEC'
                    badge = ''

                bias_color = '#E74C3C' if bias > 0 else '#2471A3'
                bias_sign = '+' if bias > 0 else ''

                summary_html += f'''
                <div style="background: {bg_color}; border: 2px solid {border_color}; border-radius: 10px; padding: 18px; text-align: center;">
                    <div style="font-size: 1.1rem; color: #1A3A5C; font-weight: 800; margin-bottom: 10px;">
                        {method} <span style="font-size: 0.75rem; color: #5D6D7E; font-weight: 600;">({ver_labels.get(method, METHOD_LABELS.get(method, chr(39)+chr(39)))})</span> {badge}
                    </div>
                    <div style="font-size: 1.6rem; font-weight: bold; color: #1A3A5C;">{avg_abs_err:.2f}%</div>
                    <div style="font-size: 0.7rem; color: #95A5A6; margin-bottom: 4px;">트리밍 평균 오차율</div>
                    <div style="font-size: 0.8rem; color: #7F8C8D;">단순평균: {np.mean(errs):.2f}%</div>
                    <div style="display: flex; justify-content: space-around; margin-top: 8px;">
                        <div>
                            <div style="font-size: 0.85rem; font-weight: 600; color: {bias_color};">{bias_sign}{bias:.1f}%</div>
                            <div style="font-size: 0.6rem; color: #ADB5BD;">편향</div>
                        </div>
                        <div>
                            <div style="font-size: 0.85rem; font-weight: 600; color: #5D6D7E;">±{std_err:.1f}%</div>
                            <div style="font-size: 0.6rem; color: #ADB5BD;">표준편차</div>
                        </div>
                        <div>
                            <div style="font-size: 0.85rem; font-weight: 600; color: #5D6D7E;">{n_stores}개</div>
                            <div style="font-size: 0.6rem; color: #ADB5BD;">매장</div>
                        </div>
                    </div>
                </div>'''
            else:
                summary_html += f'''
                <div style="background: #F8F9FA; border: 1px solid #DEE2E6; border-radius: 10px; padding: 18px; text-align: center; opacity: 0.6;">
                    <div style="font-size: 0.8rem; color: #6C757D; font-weight: 600;">{method} ({ver_labels.get(method, METHOD_LABELS.get(method, chr(39)+chr(39)))})</div>
                    <div style="font-size: 1rem; color: #ADB5BD; margin-top: 10px;">데이터 없음</div>
                </div>'''

        summary_html += '</div>'
        # 복합 방식 (v1일 때만)
        if second_row:
            summary_html += '<div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 0.5rem 0 1rem 0;">'
            for method in second_row:
                errs = [abs(r['avg_error']) for r in results_all[method]]
                if errs:
                    avg_abs_err = trimmed_mean(errs)
                    std_err = np.std(errs)
                    bias = np.mean([r['avg_error'] for r in results_all[method]])
                    n_stores = len(errs)

                    if method == best_overall:
                        border_color = '#2ECC71'
                        bg_color = '#EAFAF1'
                        badge = '<span style="background: #2ECC71; color: white; font-size: 0.6rem; padding: 2px 8px; border-radius: 10px;">★ 최적</span>'
                    elif avg_abs_err <= 12:
                        border_color = '#3498DB'
                        bg_color = '#EBF5FB'
                        badge = ''
                    elif avg_abs_err <= 18:
                        border_color = '#F39C12'
                        bg_color = '#FEF9E7'
                        badge = ''
                    else:
                        border_color = '#E74C3C'
                        bg_color = '#FDEDEC'
                        badge = ''

                    bias_color = '#E74C3C' if bias > 0 else '#2471A3'
                    bias_sign = '+' if bias > 0 else ''

                    summary_html += f'''
                <div style="background: {bg_color}; border: 2px solid {border_color}; border-radius: 10px; padding: 16px; text-align: center;">
                    <div style="font-size: 1rem; color: #1A3A5C; font-weight: 800; margin-bottom: 8px;">
                        {method} <span style="font-size: 0.7rem; color: #5D6D7E; font-weight: 600;">({ver_labels.get(method, METHOD_LABELS.get(method, chr(39)+chr(39)))})</span> {badge}
                    </div>
                    <div style="font-size: 1.4rem; font-weight: bold; color: #1A3A5C;">{avg_abs_err:.2f}%</div>
                    <div style="font-size: 0.65rem; color: #95A5A6; margin-bottom: 3px;">트리밍 평균 오차율</div>
                    <div style="font-size: 0.75rem; color: #7F8C8D;">단순평균: {np.mean(errs):.2f}%</div>
                    <div style="display: flex; justify-content: space-around; margin-top: 6px;">
                        <div>
                            <div style="font-size: 0.8rem; font-weight: 600; color: {bias_color};">{bias_sign}{bias:.1f}%</div>
                            <div style="font-size: 0.55rem; color: #ADB5BD;">편향</div>
                        </div>
                        <div>
                            <div style="font-size: 0.8rem; font-weight: 600; color: #5D6D7E;">±{std_err:.1f}%</div>
                            <div style="font-size: 0.55rem; color: #ADB5BD;">표준편차</div>
                        </div>
                        <div>
                            <div style="font-size: 0.8rem; font-weight: 600; color: #5D6D7E;">{n_stores}개</div>
                            <div style="font-size: 0.55rem; color: #ADB5BD;">매장</div>
                        </div>
                    </div>
                </div>'''
                else:
                    summary_html += f'''
                <div style="background: #F8F9FA; border: 1px solid #DEE2E6; border-radius: 10px; padding: 16px; text-align: center; opacity: 0.6;">
                    <div style="font-size: 0.8rem; color: #6C757D; font-weight: 600;">{method} ({ver_labels.get(method, METHOD_LABELS.get(method, chr(39)+chr(39)))})</div>
                    <div style="font-size: 1rem; color: #ADB5BD; margin-top: 10px;">데이터 없음</div>
                </div>'''

            summary_html += '</div>'
        st.markdown(summary_html, unsafe_allow_html=True)

        # 최적 방식 배너
        if best_overall:
            best_val = trimmed_mean([abs(r['avg_error']) for r in results_all[best_overall]])
            st.success(
                f"★ 최적 방식: **{best_overall} ({METHOD_LABELS[best_overall]})** — "
                f"트리밍 평균 오차율 {best_val:.2f}%"
            )

        # 매장별 최적 방식 카운트
        st.markdown("---")
        st.subheader("🏆 방식별 최적 매장 수 랭킹")

        # 비교할 방식 선택
        compare_methods = st.multiselect(
            "비교할 방식 선택",
            version_methods,
            default=version_methods,
            key="rank_methods"
        )

        if not compare_methods:
            compare_methods = version_methods

        best_count = {m: 0 for m in compare_methods}
        store_best_method = {}
        store_best_error = {}

        for store in all_stores:
            min_err = None
            min_method = None
            min_raw_err = None
            for method in compare_methods:
                result = validate_store(store, curve_index, method, selected_version)
                if result is not None and result['avg_error'] is not None:
                    abs_err = abs(result['avg_error'])
                    if min_err is None or abs_err < min_err:
                        min_err = abs_err
                        min_method = method
                        min_raw_err = result['avg_error']
            if min_method is not None:
                best_count[min_method] += 1
                store_best_method[store['name']] = min_method
                store_best_error[store['name']] = min_raw_err

        # 랭킹 표시 (많은 순) — HTML 바 차트
        ranked = sorted(best_count.items(), key=lambda x: x[1], reverse=True)
        total_stores = sum(best_count.values())
        max_cnt = ranked[0][1] if ranked else 1

        rank_colors = ['#2ECC71', '#3498DB', '#9B59B6', '#F39C12', '#E74C3C', '#95A5A6', '#1ABC9C', '#E67E22']
        medal = ['🥇', '🥈', '🥉', '4️⃣', '5️⃣', '6️⃣', '7️⃣', '8️⃣']

        rank_html = '<div style="margin: 1rem 0;">'
        for i, (m, cnt) in enumerate(ranked):
            pct = cnt / total_stores * 100 if total_stores > 0 else 0
            bar_width = cnt / max_cnt * 100 if max_cnt > 0 else 0
            color = rank_colors[i % len(rank_colors)]

            rank_html += f'''
            <div style="display: flex; align-items: center; margin-bottom: 10px; padding: 8px 12px; background: #FAFCFF; border-radius: 8px; border: 1px solid #E8ECF0;">
                <div style="font-size: 1.2rem; width: 36px; text-align: center;">{medal[i]}</div>
                <div style="width: 140px; font-weight: 600; color: #1A3A5C; font-size: 0.85rem;">{m} ({METHOD_LABELS[m]})</div>
                <div style="flex: 1; margin: 0 12px;">
                    <div style="background: #E8ECF0; border-radius: 6px; height: 24px; overflow: hidden;">
                        <div style="background: {color}; height: 100%; width: {bar_width}%; border-radius: 6px; transition: width 0.3s;"></div>
                    </div>
                </div>
                <div style="width: 80px; text-align: right; font-weight: 700; color: {color}; font-size: 0.95rem;">{cnt}개</div>
                <div style="width: 60px; text-align: right; color: #95A5A6; font-size: 0.8rem;">{pct:.1f}%</div>
            </div>'''

        rank_html += '</div>'
        st.markdown(rank_html, unsafe_allow_html=True)

        # 방식별 매장 목록 보기
        show_method = st.selectbox(
            "매장 목록 보기", ['선택하세요'] + [f"{m} ({METHOD_LABELS[m]})" for m in compare_methods],
            key="show_rank_stores"
        )
        if show_method != '선택하세요':
            selected_m = show_method.split(' ')[0]
            method_stores = [
                (name, err) for name, (method, err)
                in zip(store_best_method.keys(),
                       zip(store_best_method.values(), store_best_error.values()))
                if method == selected_m
            ]
            method_stores.sort(key=lambda x: abs(x[1]))

            if method_stores:
                rank_store_df = pd.DataFrame([{
                    '매장명': name,
                    '오차율': f"{err:+.2f}%"
                } for name, err in method_stores])
                st.dataframe(rank_store_df, use_container_width=True, hide_index=True)
            else:
                st.info("해당 방식이 최적인 매장이 없습니다.")

        st.markdown("---")

        # 상세 결과 방식 선택
        selected_method = st.selectbox(
            "상세 결과 방식 선택", version_methods,
            format_func=lambda x: get_method_label(x, selected_version),
            key="detail_method_sel"
        )

        # 선택된 방식 상세 테이블
        st.subheader(f"📋 방식 {selected_method} 상세 결과")
        results = results_all.get(selected_method, [])

        if results:
            tab1_data = [{
                '매장명': r['name'],
                '실제 기준 매출': f"{np.mean(r['actual']):,.0f}",
                '예측 기준 매출': f"{r['base_revenue']:,.0f}",
                '오차(액수)': f"{r['base_revenue'] - np.mean(r['actual']):+,.0f}",
                '오차율': f"{r['avg_error']:+.2f}%",
                '최적방식': store_best_method.get(r['name'], '-')
            } for r in sorted(results, key=lambda x: abs(x['avg_error']))]

            headers_t1 = ['매장명', '실제 기준 매출', '예측 기준 매출', '오차(액수)', '오차율', '최적방식']
            t1_html = '<div style="max-height: 500px; overflow-y: auto;"><table style="width:100%; border-collapse: collapse; font-size: 0.85rem;">'
            t1_html += '<thead><tr>'
            for h in headers_t1:
                t1_html += f'<th style="background: #1A3A5C; color: white; font-weight: bold; text-align: center; padding: 10px 8px; position: sticky; top: 0;">{h}</th>'
            t1_html += '</tr></thead><tbody>'
            for i, d in enumerate(tab1_data):
                bg = '#F8FBFF' if i % 2 == 0 else '#FFFFFF'
                t1_html += f'<tr style="background: {bg};">'
                for h in headers_t1:
                    val = d.get(h, '')
                    t1_html += f'<td style="text-align: center; padding: 7px 8px; border-bottom: 1px solid #E8ECF0;">{val}</td>'
                t1_html += '</tr>'
            t1_html += '</tbody></table></div>'
            st.markdown(t1_html, unsafe_allow_html=True)

            # 분포 차트
            st.subheader("📊 오차율 분포")
            errors_list = [r['avg_error'] for r in results]
            fig_hist = go.Figure()
            fig_hist.add_trace(go.Histogram(
                x=errors_list, nbinsx=20,
                marker_color='#2471A3', opacity=0.8
            ))
            fig_hist.update_layout(
                xaxis_title="오차율 (%)",
                yaxis_title="매장 수",
                template="plotly_white",
                height=350
            )
            st.plotly_chart(fig_hist, use_container_width=True)
        else:
            st.warning("검증 가능한 매장이 없습니다.")

    # ═══════════════════════════════════════════════════════════════
    # 탭 2: 개별 매장 조회
    # ═══════════════════════════════════════════════════════════════
    with tab2:
        st.subheader("🔍 개별 매장 상세 조회")

        store_names = [s['name'] for s in all_stores]
        ver_end_map = {'v1': 3, 'v2': 4, 'v3': 5, 'v4': 6, 'v5': 7, 'v6': 8}
        col_store, col_ver, col_method = st.columns([3, 2, 2])
        with col_store:
            selected_store_name = st.selectbox(
                "매장 선택", store_names,
                index=0 if store_names else None
            )
        with col_ver:
            selected_version_tab2 = st.selectbox(
                "버전 선택",
                ['v1', 'v2', 'v3', 'v4', 'v5', 'v6'],
                format_func=lambda x: f"V{x[1]}: m1~m{ver_end_map[x]} 데이터",
                key="tab2_version"
            )
        with col_method:
            start_m_map = {'A': 1, 'B': 2, 'C': 3}
            def fmt_method_tab2(x):
                if x in ('AB', 'AC', 'BC', 'ABC'):
                    return f"{x}: {METHOD_LABELS[x]}"
                s = start_m_map[x]
                e = ver_end_map.get(selected_version_tab2, 3)
                if s == e:
                    return f"{x}: m{s} 사용"
                return f"{x}: m{s}~m{e} 사용"
            selected_method_tab2 = st.selectbox(
                "방식 선택", ALL_METHODS,
                format_func=fmt_method_tab2,
                key="tab2_method"
            )

        if selected_store_name:
            store = next(s for s in all_stores if s['name'] == selected_store_name)

            # 6가지 방식 결과
            st.markdown(f"### 📌 {selected_store_name}")

            # 데이터 월차 정보 표시
            max_month = get_store_max_month(store)
            st.markdown(f"**데이터: m0 ~ m{max_month} ({max_month+1}개월)**")

            # 실제평균 한 번만 표시 (항상 m4~m9 기준, 버전 무관)
            sample_result = validate_store(store, curve_index, 'A', 'v1')
            if sample_result and sample_result['actual'] is not None:
                actual_avg = np.mean(sample_result['actual'])
                st.markdown(f"**실제 기준 매출(오픈4~9개월 평균)=100 : :blue[{actual_avg:,.0f}원]**")
            elif max_month < 9:
                st.caption("⚠️ m9까지 데이터가 없어 오차율 검증 불가, 예측 기준 매출만 표시합니다.")

            # 6가지 방식을 HTML 카드 형식으로
            method_results = []
            tab2_end = ver_end_map.get(selected_version_tab2, 3)
            for method in ALL_METHODS:
                # 동적 라벨
                s = {'A': 1, 'B': 2, 'C': 3}.get(method)
                if s and s <= tab2_end:
                    label = f"m{s}~m{tab2_end} 사용" if s != tab2_end else f"m{s} 사용"
                else:
                    label = METHOD_LABELS.get(method, '')
                result = validate_store(store, curve_index, method, selected_version_tab2)
                if result:
                    method_results.append({
                        'method': method,
                        'label': label,
                        'base': result['base_revenue'],
                        'error': result['avg_error'],
                        'available': True
                    })
                else:
                    method_results.append({
                        'method': method,
                        'label': label,
                        'base': 0,
                        'error': None,
                        'available': False
                    })

            # 최적 방식 찾기 (오차율이 있는 것 중에서)
            available_with_error = [r for r in method_results if r['available'] and r['error'] is not None]
            available = [r for r in method_results if r['available']]
            best_method = min(available_with_error, key=lambda x: abs(x['error']))['method'] if available_with_error else None

            # HTML 카드 그리드 (3+4 배치)
            cards_html = '<div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin: 1rem 0;">'
            for r in [mr for mr in method_results if mr['method'] in ('A', 'B', 'C')]:
                if not r['available']:
                    cards_html += f'''
                    <div style="background: #F8F9FA; border: 1px solid #DEE2E6; border-radius: 8px; padding: 16px; text-align: center; opacity: 0.6;">
                        <div style="font-size: 0.75rem; color: #6C757D; margin-bottom: 4px;">{r['method']} ({r['label']})</div>
                        <div style="font-size: 0.9rem; color: #ADB5BD;">데이터 부족</div>
                    </div>'''
                else:
                    # 색상 결정
                    abs_err = abs(r['error']) if r['error'] is not None else None
                    if r['method'] == best_method and abs_err is not None:
                        border_color = '#2ECC71'
                        bg_color = '#EAFAF1'
                        badge = '<span style="background: #2ECC71; color: white; font-size: 0.65rem; padding: 2px 6px; border-radius: 10px; margin-left: 4px;">최적</span>'
                    elif abs_err is not None and abs_err <= 10:
                        border_color = '#3498DB'
                        bg_color = '#EBF5FB'
                        badge = ''
                    elif abs_err is not None and abs_err <= 20:
                        border_color = '#F39C12'
                        bg_color = '#FEF9E7'
                        badge = ''
                    elif abs_err is not None:
                        border_color = '#E74C3C'
                        bg_color = '#FDEDEC'
                        badge = ''
                    else:
                        border_color = '#95A5A6'
                        bg_color = '#F8F9FA'
                        badge = ''

                    if r['error'] is not None:
                        err_color = '#E74C3C' if r['error'] > 0 else '#2471A3'
                        sign = '+' if r['error'] > 0 else ''
                        error_display = f'<div style="font-size: 1rem; font-weight: bold; color: {err_color}; margin-top: 6px;">{sign}{r["error"]:.2f}%</div><div style="font-size: 0.7rem; color: #95A5A6;">오차율</div>'
                    else:
                        error_display = '<div style="font-size: 0.8rem; color: #95A5A6; margin-top: 6px;">오차율 산정 불가</div>'

                    cards_html += f'''
                    <div style="background: {bg_color}; border: 2px solid {border_color}; border-radius: 8px; padding: 16px; text-align: center;">
                        <div style="font-size: 0.75rem; color: #5D6D7E; margin-bottom: 8px; font-weight: 600;">{r['method']} ({r['label']}){badge}</div>
                        <div style="font-size: 1.1rem; font-weight: bold; color: #1A3A5C;">{r['base']:,.0f}원</div>
                        <div style="font-size: 0.7rem; color: #95A5A6; margin: 2px 0;">예측 기준 매출</div>
                        {error_display}
                    </div>'''

            cards_html += '</div>'
            # 복합 방식 (4열)
            cards_html += '<div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 0.5rem 0 1rem 0;">'
            for r in [mr for mr in method_results if mr['method'] in ('AB', 'AC', 'BC', 'ABC')]:
                if not r['available']:
                    cards_html += f'''
                    <div style="background: #F8F9FA; border: 1px solid #DEE2E6; border-radius: 8px; padding: 14px; text-align: center; opacity: 0.6;">
                        <div style="font-size: 0.75rem; color: #6C757D; margin-bottom: 4px;">{r['method']} ({r['label']})</div>
                        <div style="font-size: 0.9rem; color: #ADB5BD;">데이터 부족</div>
                    </div>'''
                else:
                    abs_err = abs(r['error']) if r['error'] is not None else None
                    if r['method'] == best_method and abs_err is not None:
                        border_color = '#2ECC71'
                        bg_color = '#EAFAF1'
                        badge = '<span style="background: #2ECC71; color: white; font-size: 0.65rem; padding: 2px 6px; border-radius: 10px; margin-left: 4px;">최적</span>'
                    elif abs_err is not None and abs_err <= 10:
                        border_color = '#3498DB'
                        bg_color = '#EBF5FB'
                        badge = ''
                    elif abs_err is not None and abs_err <= 20:
                        border_color = '#F39C12'
                        bg_color = '#FEF9E7'
                        badge = ''
                    elif abs_err is not None:
                        border_color = '#E74C3C'
                        bg_color = '#FDEDEC'
                        badge = ''
                    else:
                        border_color = '#95A5A6'
                        bg_color = '#F8F9FA'
                        badge = ''

                    if r['error'] is not None:
                        err_color = '#E74C3C' if r['error'] > 0 else '#2471A3'
                        sign = '+' if r['error'] > 0 else ''
                        error_display = f'<div style="font-size: 0.9rem; font-weight: bold; color: {err_color}; margin-top: 4px;">{sign}{r["error"]:.2f}%</div><div style="font-size: 0.65rem; color: #95A5A6;">오차율</div>'
                    else:
                        error_display = '<div style="font-size: 0.75rem; color: #95A5A6; margin-top: 4px;">오차율 산정 불가</div>'

                    cards_html += f'''
                    <div style="background: {bg_color}; border: 2px solid {border_color}; border-radius: 8px; padding: 14px; text-align: center;">
                        <div style="font-size: 0.7rem; color: #5D6D7E; margin-bottom: 6px; font-weight: 600;">{r['method']} ({r['label']}){badge}</div>
                        <div style="font-size: 1rem; font-weight: bold; color: #1A3A5C;">{r['base']:,.0f}원</div>
                        <div style="font-size: 0.65rem; color: #95A5A6; margin: 2px 0;">예측 기준 매출</div>
                        {error_display}
                    </div>'''

            cards_html += '</div>'
            st.markdown(cards_html, unsafe_allow_html=True)

            # 선택 방식 상세
            result = validate_store(store, curve_index, selected_method_tab2, selected_version_tab2)
            if result and result['actual'] is not None:
                st.markdown("---")
                st.markdown(f"#### 방식 {selected_method_tab2} 상세")

                v_start = result['verify_start']
                v_end = result['verify_end']
                detail_rows = list(zip(
                    [f'm{i} ({curve_index.get(i, 0):.1f}%)' for i in range(v_start, v_end)],
                    [f"{a:,.0f}" for a in result['actual']],
                    [f"{p:,.0f}" for p in result['predicted']],
                    [f"{e:+.1f}%" for e in result['errors']]
                ))
                det_headers = ['월차', '실제 매출', '예측 매출', '오차율(%)']
                det_html = '<table style="width:100%; border-collapse: collapse; font-size: 0.85rem;">'
                det_html += '<thead><tr>'
                for h in det_headers:
                    det_html += f'<th style="background: #1A3A5C; color: white; font-weight: bold; text-align: center; padding: 10px 8px;">{h}</th>'
                det_html += '</tr></thead><tbody>'
                for i, row in enumerate(detail_rows):
                    bg = '#F8FBFF' if i % 2 == 0 else '#FFFFFF'
                    det_html += f'<tr style="background: {bg};">'
                    for val in row:
                        det_html += f'<td style="text-align: center; padding: 7px 8px; border-bottom: 1px solid #E8ECF0;">{val}</td>'
                    det_html += '</tr>'
                det_html += '</tbody></table>'
                st.markdown(det_html, unsafe_allow_html=True)

                # 예측 vs 실제 비교 차트
                fig = go.Figure()
                months = [f'm{i}' for i in range(v_start, v_end)]
                fig.add_trace(go.Bar(
                    x=months, y=result['actual'],
                    name='실제매출', marker_color='#2471A3'
                ))
                fig.add_trace(go.Bar(
                    x=months, y=result['predicted'],
                    name='예측매출', marker_color='#E74C3C', opacity=0.7
                ))
                fig.update_layout(
                    title=f"{selected_store_name} — 실제 vs 예측",
                    xaxis_title="월차", yaxis_title="매출 (원)",
                    barmode='group', template="plotly_white",
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)

            elif result and result['base_revenue'] is not None:
                # m4~m9 부족하지만 기준매출은 있는 경우: 있는 데이터만 비교
                st.markdown("---")
                st.markdown(f"#### 방식 {selected_method_tab2} 상세 (보유 데이터 기준)")
                sales = store['sales']
                max_m = get_store_max_month(store)

                # 보유한 월차 중 m1 이후로 예측 가능한 범위
                detail_months = []
                detail_actual = []
                detail_predicted = []
                detail_errors = []
                for m in range(1, max_m + 1):
                    if m < len(sales) and sales[m] is not None and sales[m] > 0 and m in curve_index:
                        pred = result['base_revenue'] * (curve_index[m] / 100)
                        actual = sales[m]
                        err = (pred - actual) / actual * 100
                        detail_months.append(f'm{m}')
                        detail_actual.append(actual)
                        detail_predicted.append(pred)
                        detail_errors.append(err)

                if detail_months:
                    detail_rows2 = list(zip(
                        [f'{m} ({curve_index.get(int(m[1:]), 0):.1f}%)' for m in detail_months],
                        [f"{a:,.0f}" for a in detail_actual],
                        [f"{p:,.0f}" for p in detail_predicted],
                        [f"{e:+.1f}%" for e in detail_errors]
                    ))
                    det_headers2 = ['월차', '실제 매출', '예측 매출', '오차율(%)']
                    det_html2 = '<table style="width:100%; border-collapse: collapse; font-size: 0.85rem;">'
                    det_html2 += '<thead><tr>'
                    for h in det_headers2:
                        det_html2 += f'<th style="background: #1A3A5C; color: white; font-weight: bold; text-align: center; padding: 10px 8px;">{h}</th>'
                    det_html2 += '</tr></thead><tbody>'
                    for i, row in enumerate(detail_rows2):
                        bg = '#F8FBFF' if i % 2 == 0 else '#FFFFFF'
                        det_html2 += f'<tr style="background: {bg};">'
                        for val in row:
                            det_html2 += f'<td style="text-align: center; padding: 7px 8px; border-bottom: 1px solid #E8ECF0;">{val}</td>'
                        det_html2 += '</tr>'
                    det_html2 += '</tbody></table>'
                    st.markdown(det_html2, unsafe_allow_html=True)

                    # 차트
                    fig = go.Figure()
                    fig.add_trace(go.Bar(
                        x=detail_months, y=detail_actual,
                        name='실제매출', marker_color='#2471A3'
                    ))
                    fig.add_trace(go.Bar(
                        x=detail_months, y=detail_predicted,
                        name='예측매출', marker_color='#E74C3C', opacity=0.7
                    ))
                    fig.update_layout(
                        title=f"{selected_store_name} — 실제 vs 예측 (보유 데이터)",
                        xaxis_title="월차", yaxis_title="매출 (원)",
                        barmode='group', template="plotly_white",
                        height=400
                    )
                    st.plotly_chart(fig, use_container_width=True)

            # 전체 매출 추이
            st.markdown("---")
            st.markdown("#### 📈 전체 월별 매출 추이")
            sales = store['sales']
            valid_sales = [(i, s) for i, s in enumerate(sales)
                          if s is not None and s > 0]
            if valid_sales:
                fig2 = go.Figure()
                x_vals = [f'm{i}' for i, _ in valid_sales]
                y_vals = [s for _, s in valid_sales]
                fig2.add_trace(go.Scatter(
                    x=x_vals, y=y_vals,
                    mode='lines+markers',
                    name='실제매출',
                    line=dict(color='#1A3A5C', width=2),
                    marker=dict(size=8)
                ))

                # 성장곡선 예측 오버레이
                result_for_curve = validate_store(store, curve_index, selected_method_tab2, selected_version_tab2)
                if result_for_curve:
                    preds = predict_growth_curve(
                        result_for_curve['base_revenue'], curve_index
                    )
                    pred_months = [f'm{m}' for m in sorted(preds.keys())]
                    pred_vals = [preds[m] for m in sorted(preds.keys())]
                    fig2.add_trace(go.Scatter(
                        x=pred_months, y=pred_vals,
                        mode='lines',
                        name='예측 매출',
                        line=dict(color='#E74C3C', width=2, dash='dash')
                    ))

                # 회귀밴드 (타겟수요 기반)
                target_data = load_target_data()
                target_val = match_target(selected_store_name, target_data)
                if target_val:
                    band = get_regression_band(target_val)
                    fig2.add_hline(y=band['predicted'], line_dash="solid",
                                   line_color="#2ECC71", opacity=0.8,
                                   annotation_text=f"예상매출 {band['predicted']:,.0f}")
                    fig2.add_hline(y=band['upper'], line_dash="dash",
                                   line_color="#3498DB", opacity=0.6,
                                   annotation_text=f"상한 {band['upper']:,.0f}")
                    fig2.add_hline(y=band['lower'], line_dash="dash",
                                   line_color="#E74C3C", opacity=0.6,
                                   annotation_text=f"하한 {band['lower']:,.0f}")
                    # 밴드 영역 음영
                    fig2.add_hrect(y0=band['lower'], y1=band['upper'],
                                   fillcolor="#2ECC71", opacity=0.07,
                                   line_width=0)

                fig2.update_layout(
                    title=f"{selected_store_name} — 월별 매출 추이"
                          + (f" (타겟수요: {int(target_val)})" if target_val else ""),
                    xaxis_title="월차", yaxis_title="매출 (원)",
                    template="plotly_white", height=450
                )
                st.plotly_chart(fig2, use_container_width=True)

    # ═══════════════════════════════════════════════════════════════
    # 탭 5: 그룹별 분산 분석
    # ═══════════════════════════════════════════════════════════════
    with tab5:
        st.subheader("📉 그룹 별 곡선지수 표준편차 분석")

        # 데이터 로드
        @st.cache_data
        def load_store_curves():
            wb = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)
            ws = wb['Calc_지수곡선(시즌미반영)']
            stores_curve = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[0]
                group = str(row[3]).strip() if row[3] else None
                if name is None or group is None or group == '3+':
                    continue
                indices = []
                for i in range(6, min(54, len(row))):
                    indices.append(row[i] if row[i] is not None else None)
                stores_curve.append({'name': str(name), 'group': group, 'indices': indices})
            wb.close()
            return stores_curve

        stores_curve = load_store_curves()

        # ── 4개 그룹 가중평균 표준편차 (맨 위) ──
        def calc_weighted_std(store_list):
            stds_list, counts_list = [], []
            for m_idx in range(48):
                vals = [s['indices'][m_idx] for s in store_list
                        if m_idx < len(s['indices']) and s['indices'][m_idx] is not None]
                if len(vals) >= 2:
                    stds_list.append(np.std(vals, ddof=1))
                    counts_list.append(len(vals))
            return np.average(stds_list, weights=counts_list) if stds_list else 0

        g0 = [s for s in stores_curve if s['group'] == '0']
        g1 = [s for s in stores_curve if s['group'] == '1']
        g2 = [s for s in stores_curve if s['group'] == '2']

        std_g0 = calc_weighted_std(g0)
        std_g1 = calc_weighted_std(g1)
        std_g2 = calc_weighted_std(g2)
        std_all = calc_weighted_std(stores_curve)

        st.markdown(f'''
        <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 1rem 0;">
            <div style="background: #EBF5FB; border: 2px solid #2471A3; border-radius: 10px; padding: 18px; text-align: center;">
                <div style="font-size: 0.8rem; color: #5D6D7E; font-weight: 600;">그룹 0 (경쟁 0)</div>
                <div style="font-size: 1.8rem; font-weight: bold; color: #1A3A5C;">{std_g0:.2f}</div>
                <div style="font-size: 0.65rem; color: #95A5A6;">{len(g0)}개 매장</div>
            </div>
            <div style="background: #EAFAF1; border: 2px solid #2ECC71; border-radius: 10px; padding: 18px; text-align: center;">
                <div style="font-size: 0.8rem; color: #5D6D7E; font-weight: 600;">그룹 1</div>
                <div style="font-size: 1.8rem; font-weight: bold; color: #1E8449;">{std_g1:.2f}</div>
                <div style="font-size: 0.65rem; color: #95A5A6;">{len(g1)}개 매장</div>
            </div>
            <div style="background: #FEF9E7; border: 2px solid #F39C12; border-radius: 10px; padding: 18px; text-align: center;">
                <div style="font-size: 0.8rem; color: #5D6D7E; font-weight: 600;">그룹 2</div>
                <div style="font-size: 1.8rem; font-weight: bold; color: #D68910;">{std_g2:.2f}</div>
                <div style="font-size: 0.65rem; color: #95A5A6;">{len(g2)}개 매장</div>
            </div>
            <div style="background: #F4ECF7; border: 2px solid #8E44AD; border-radius: 10px; padding: 18px; text-align: center;">
                <div style="font-size: 0.8rem; color: #5D6D7E; font-weight: 600;">전체 (0~2)</div>
                <div style="font-size: 1.8rem; font-weight: bold; color: #6C3483;">{std_all:.2f}</div>
                <div style="font-size: 0.65rem; color: #95A5A6;">{len(stores_curve)}개 매장</div>
            </div>
        </div>
        <div style="text-align: center; font-size: 0.7rem; color: #95A5A6; margin-bottom: 1rem;">m1~m48 가중평균 표준편차 (매장수 가중)</div>
        ''', unsafe_allow_html=True)

        # ── 그룹 선택 ──
        groups_available = sorted(set(s['group'] for s in stores_curve))
        selected_group = st.selectbox("그룹 선택", ['전체(0~2)'] + groups_available, key="group_sel")

        if selected_group == '전체(0~2)':
            filtered = stores_curve
        else:
            filtered = [s for s in stores_curve if s['group'] == selected_group]

        st.caption(f"대상 매장 수: {len(filtered)}개")

        if filtered:
            max_months = 48
            stats_data = []
            for m_idx in range(max_months):
                vals = [s['indices'][m_idx] for s in filtered
                        if m_idx < len(s['indices']) and s['indices'][m_idx] is not None]
                if len(vals) >= 2:
                    avg = np.mean(vals)
                    std = np.std(vals, ddof=1)
                    var = np.var(vals, ddof=1)
                    cv = (std / avg * 100) if avg != 0 else 0
                    w_avg = curve_index.get(m_idx + 1, None)
                    stats_data.append({
                        '월차': f'm{m_idx+1}', '매장수': len(vals),
                        '가중평균지수': round(w_avg, 2) if w_avg else '-',
                        '그룹평균': round(avg, 2),
                        '표준편차(분산)': f"{std:.2f} ({var:.2f})",
                        'CV(%)': round(cv, 2),
                        '_std': std, '_avg': avg
                    })

            if stats_data:
                months_list = [d['월차'] for d in stats_data]
                avgs = [d['_avg'] for d in stats_data]
                stds = [d['_std'] for d in stats_data]

                fig_var = go.Figure()
                fig_var.add_trace(go.Scatter(x=months_list, y=avgs, mode='lines+markers', name='그룹 평균', line=dict(color='#1A3A5C', width=2)))
                fig_var.add_trace(go.Scatter(x=months_list, y=[a+s for a,s in zip(avgs,stds)], mode='lines', name='+1σ', line=dict(color='#3498DB', width=1, dash='dash')))
                fig_var.add_trace(go.Scatter(x=months_list, y=[a-s for a,s in zip(avgs,stds)], mode='lines', name='-1σ', line=dict(color='#E74C3C', width=1, dash='dash'), fill='tonexty', fillcolor='rgba(52,152,219,0.1)'))
                fig_var.add_hline(y=100, line_dash="dot", line_color="gray", annotation_text="기준(100)")
                fig_var.update_layout(title=f"그룹 {'전체(0~2)' if selected_group=='전체(0~2)' else selected_group} — 곡선지수 평균 ± 표준편차", xaxis_title="월차", yaxis_title="곡선지수(%)", template="plotly_white", height=450)
                st.plotly_chart(fig_var, use_container_width=True)

                with st.expander("📋 월차별 상세 데이터 보기"):
                    display_data = [{k: v for k, v in d.items() if not k.startswith('_')} for d in stats_data]
                    # HTML 테이블로 렌더링
                    headers = ['월차', '매장수', '가중평균지수', '그룹평균', '표준편차(분산)', 'CV(%)']
                    table_html = '<div style="max-height: 400px; overflow-y: auto;"><table style="width:100%; border-collapse: collapse; font-size: 0.85rem;">'
                    table_html += '<thead><tr>'
                    for h in headers:
                        table_html += f'<th style="background: #1A3A5C; color: white; font-weight: bold; text-align: center; padding: 10px 8px; position: sticky; top: 0;">{h}</th>'
                    table_html += '</tr></thead><tbody>'
                    for i, d in enumerate(display_data):
                        bg = '#F8FBFF' if i % 2 == 0 else '#FFFFFF'
                        table_html += f'<tr style="background: {bg};">'
                        for h in headers:
                            val = d.get(h, '')
                            if h == 'CV(%)' and isinstance(val, (int, float)):
                                bar_w = min(val / 30 * 100, 100)
                                color = '#2ECC71' if val < 10 else '#F39C12' if val < 20 else '#E74C3C'
                                table_html += f'<td style="text-align: center; padding: 7px 8px; border-bottom: 1px solid #E8ECF0;"><div style="display:flex; align-items:center; gap:6px;"><div style="flex:1; background:#E8ECF0; border-radius:4px; height:14px; overflow:hidden;"><div style="background:{color}; height:100%; width:{bar_w}%; border-radius:4px;"></div></div><span style="font-size:0.75rem; min-width:40px;">{val:.1f}%</span></div></td>'
                            else:
                                table_html += f'<td style="text-align: center; padding: 7px 8px; border-bottom: 1px solid #E8ECF0;">{val}</td>'
                        table_html += '</tr>'
                    table_html += '</tbody></table></div>'
                    st.markdown(table_html, unsafe_allow_html=True)

                st.markdown("---")
                st.subheader("🏪 개별 매장 곡선지수")
                show_stores = st.multiselect("매장 선택 (최대 10개)", [s['name'] for s in filtered], max_selections=10, key="var_stores")
                if show_stores:
                    group_info = ", ".join([f"{s['name'].split(') ')[1] if ')' in s['name'] else s['name']} (그룹{s['group']})" for s in filtered if s['name'] in show_stores])
                    st.caption(f"📌 {group_info}")
                    fig_ind = go.Figure()
                    fig_ind.add_trace(go.Scatter(x=months_list, y=avgs, mode='lines', name='그룹 평균', line=dict(color='#95A5A6', width=3, dash='dot')))
                    colors = ['#1A3A5C','#E74C3C','#2ECC71','#9B59B6','#F39C12','#1ABC9C','#E67E22','#8E44AD','#2980B9','#27AE60']
                    for idx, sname in enumerate(show_stores):
                        s = next(x for x in filtered if x['name'] == sname)
                        vals = [v for v in s['indices'][:max_months] if v is not None]
                        fig_ind.add_trace(go.Scatter(x=[f'm{i+1}' for i in range(len(vals))], y=vals, mode='lines+markers', name=sname, line=dict(color=colors[idx%len(colors)], width=2), marker=dict(size=5)))
                    fig_ind.update_layout(title="개별 매장 곡선지수 vs 그룹 평균", xaxis_title="월차", yaxis_title="곡선지수(%)", template="plotly_white", height=450)
                    st.plotly_chart(fig_ind, use_container_width=True)


if __name__ == "__main__":
    main()
